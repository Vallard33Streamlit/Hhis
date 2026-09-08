import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import streamlit_antd_components as sac
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment


st.set_page_config(layout="wide")

df = pd.read_csv(
    "hhis_et_autres.csv",
    dtype={"Code HS6": str, "Code HS4": str, "Code ISIC": str, "Code HS6 2017": str},
)


@st.cache_data
def load_labels():
    labels_sections = pd.read_csv("labels_sections.csv")
    labels_sections["l_hs2"] = labels_sections["l_hs2"].fillna("").apply(lambda x: x.split(","))
    labels_sections.loc[labels_sections["Niveau"] == "HS2", "l_hs2"] = np.nan
    labels_sections_tree = []
    for s in labels_sections[labels_sections["Niveau"] == "Section"].iterrows():
        dic = {}
        dic["label"] = "Section " + s[1].Catégorie + " - " + s[1].Label
        dic["value"] = s[1].Catégorie
        children = []
        for c in labels_sections[
            labels_sections["Catégorie"].apply(lambda x: x in s[1].l_hs2)
        ].iterrows():
            dic_c = {}
            dic_c["label"] = c[1].Catégorie + " - " + c[1].Label
            dic_c["value"] = c[1].Catégorie
            children.append(dic_c)
        dic["children"] = children
        labels_sections_tree.append(dic)
    return labels_sections, labels_sections_tree


# --- Sidebar : Filtres globaux ---
st.sidebar.header("Filtres globaux")
l_filtres = []

log_values = np.linspace(-3, 6, 901)
filter_v_imp_fr = st.sidebar.checkbox(
    "Filtrer les produits grâce aux importations européennes 2024",
    key="filter_v_imp_fr",
)
if filter_v_imp_fr:
    log_value_v_imp_fr = st.sidebar.select_slider(
        "Importations européennes (en 1000$) supérieur à :",
        options=log_values,
        value=-3,
        format_func=lambda x: f"{10**x:.3f}",
    )
    df = df[df["Importations"] >= 10**log_value_v_imp_fr]
    l_filtres.append(
        ("Importations européennes (en 1000$) supérieur à :", 10**log_value_v_imp_fr)
    )


labels_sections, labels_sections_tree = load_labels()
# Filtre hs2
with st.sidebar.expander(
    "Sélectionner les sections ou catégories HS2 à afficher",
    expanded=False,
    key="container_labels_sections_tree",
):
    selected_categories_index = sac.tree(
        labels_sections_tree,
        open_index=[],
        index=list(labels_sections.index),
        checkbox=True,
        return_index=True,
        key="labels_sections_tree",
        height=500,
    )
    selected_categories = labels_sections.loc[
        selected_categories_index, "Catégorie"
    ].values
    df = df[df["Code HS6"].apply(lambda x: x[:2] in selected_categories)]

# Filtre taux_couverture_imp_exp
filter_by_taux_couverture = st.sidebar.checkbox(
    "Filtrer les produits grâce à un taux de couverture",
    key="filter_by_taux_couverture",
)
if filter_by_taux_couverture:
    taux_couverture = st.sidebar.slider(
        "Rapport entre les exportations et les importations inférieur à :",
        min_value=0.0,
        max_value=5.0,
        value=1.0,
        step=0.1,
        format="%.2f",
    )
    df = df[df["Taux de couverture Exportations/Importations"] <= taux_couverture]
    l_filtres.append(
        ("Rapport entre les exportations et les importations inférieur à :", taux_couverture)
    )

# Filtre taux_couverture_transit
filter_by_taux_couverture_transit = st.sidebar.checkbox(
    "Filtrer les produits grâce à un taux de transit",
    key="filter_by_taux_couverture_transit",
)
if filter_by_taux_couverture_transit:
    rap_flux_transit = st.sidebar.slider(
        "Rapport entre les flux transitants en Europe et les importations inférieur à :",
        min_value=0.0,
        max_value=5.0,
        value=1.0,
        step=0.1,
        format="%.2f",
    )
    df = df[df["Taux de couverture de transit (Transit/Importations)"] <= rap_flux_transit]
    l_filtres.append(
        (
            "Rapport entre les flux transitants en Europe et les importations inférieur à :",
            rap_flux_transit,
        )
    )

# Filtre taux_couverture_imp_exp
filter_by_taux_couverture_fr = st.sidebar.checkbox(
    "Filtrer les produits grâce à un taux de couverture français",
    key="filter_by_taux_couverture_fr",
)
if filter_by_taux_couverture_fr:
    taux_couverture_fr = st.sidebar.slider(
        "Rapport entre les exportations et les importations françaises inférieur à :",
        min_value=0.0,
        max_value=5.0,
        value=1.0,
        step=0.1,
        format="%.2f",
    )
    df = df[
        df["Taux de couverture Exportations/Importations françaises"]
        <= taux_couverture_fr
    ]
    l_filtres.append(
        (
            "Rapport entre les exportations et les importations françaises inférieur à :",
            taux_couverture_fr,
        )
    )

# Filtre sur éléments équivalent trendeo
eq_trendeo_checkbox = st.sidebar.checkbox(
    "Garder uniquement les produits qui ont une nomenclature trendeo",
    key="eq_trendeo_checkbox",
)
if eq_trendeo_checkbox:
    df = df[df["Label ISIC"].notna()]
    l_filtres.append(
        ("Garder uniquement les produits qui ont une nomenclature trendeo", True)
    )

# Filtre sur les premiers pays
filter_by_countries = st.sidebar.checkbox(
    "Filtrer les produits par pays exportateurs", key="filter_by_countries"
)
if filter_by_countries:
    st.sidebar.markdown(
        '<p style="font-size: 14px; margin: 0 0 0.5rem 0;">Prendre les <b>n₁</b> premiers exportateurs vers l\'Europe parmi les <b>n₂</b> plus gros exportateurs mondiaux</p>',
        unsafe_allow_html=True,
    )
    col1, col2 = st.sidebar.columns(2)
    n1 = col1.selectbox("n₁", range(1, 6), key="n1")
    n2 = col2.selectbox("n₂", range(n1, 6), key="n2")
    rangs = ["Premier", "Deuxième", "Troisième", "Quatrième", "Cinquième"]
    if n1 != 0:
        s_exp_ue = df[[f"{rang} exporateur vers l'UE" for rang in rangs[:n1]]]
        s_exp_m = df[[f"{rang} exportateur mondial" for rang in rangs[:n2]]]
        mask = s_exp_ue.apply(frozenset, axis=1).combine(
            s_exp_m.apply(frozenset, axis=1),
            lambda a, b: len(a & b) == len(a),
        )
        df = df[mask]
    l_filtres.append(
        (
            "Prendre les n₁ premiers exportateurs vers l'Europe parmi les n₂ plus gros exportateurs mondiaux",
            f"n₁ = {n1}, n₂ = {n2}",
        )
    )

n_ess_elasticite = df["Essentialité grâce aux élasticités"].sum()

# Filtre hhi_EU
st.header("Différents filtres sur les indices HHi et les parts d'investissements")
filter_by_hhi_eu = st.checkbox(
    "Filtrer les produits par l'indice HHi européen", key="filter_by_hhi_eu"
)
if filter_by_hhi_eu:
    hhi_eu = st.slider(
        "HHi européen supérieur à :",
        min_value=0.0,
        max_value=1.0,
        value=0.25,
        step=0.01,
        format="%.2f",
    )
    df = df[df["HHi Européen"] >= hhi_eu]
    l_filtres.append(("HHi européen supérieur à :", hhi_eu))

# Filtre hhi_M
filter_by_hhi_m = st.checkbox(
    "Filtrer les produits par l'indice HHi mondial", key="filter_by_hhi_m"
)
if filter_by_hhi_m:
    hhi_M = st.slider(
        "HHi mondial supérieur à :",
        min_value=0.0,
        max_value=1.0,
        value=0.25,
        step=0.01,
        format="%.2f",
    )
    df = df[df["HHi Mondial"] >= hhi_M]
    l_filtres.append(("HHi mondial supérieur à :", hhi_M))

# Filtre hhi_FR
filter_by_hhi_fr = st.checkbox(
    "Filtrer les produits par l'indice HHi français", key="filter_by_hhi_fr"
)
if filter_by_hhi_fr:
    hhi_FR = st.slider(
        "HHi français supérieur à :",
        min_value=0.0,
        max_value=1.0,
        value=0.25,
        step=0.01,
        format="%.2f",
    )
    df = df[df["HHi Français"] >= hhi_FR]
    l_filtres.append(("HHi Français supérieur à :", hhi_FR))

# Filtre part_UE_CDV
filter_by_part_UE_CDV = st.checkbox(
    "Filtrer les produits par la part européenne du contrôle opérationel",
    key="filter_by_part_UE_CDV",
)
if filter_by_part_UE_CDV:
    part_UE_CDV = st.slider(
        "Part UE Contrôle opérationel inférieure à :",
        min_value=0.0,
        max_value=1.0,
        value=0.1,
        step=0.01,
        format="%.2f",
    )
    df = df[
        (df["Part UE Contrôle opérationel"] <= part_UE_CDV)
        | (df["Part UE Contrôle opérationel"].isna())
    ]
    l_filtres.append(("Part UE Contrôle opérationel inférieure à :", part_UE_CDV))

# Filtre part_UE_CF
filter_by_part_UE_CF = st.checkbox(
    "Filtrer les produits par la part européenne du contrôle financier",
    key="filter_by_part_UE_CF",
)
if filter_by_part_UE_CF:
    part_UE_CF = st.slider(
        "Part UE Contrôle financier inférieure à :",
        min_value=0.0,
        max_value=1.0,
        value=0.1,
        step=0.01,
        format="%.2f",
    )
    df = df[
        (df["Part UE Contrôle financier"] <= part_UE_CF)
        | (df["Part UE Contrôle financier"].isna())
    ]
    l_filtres.append(("Part UE Contrôle financier inférieure à :", part_UE_CF))

# Filtre hhi_CDV
filter_by_hhi_CDV = st.checkbox(
    "Filtrer les produits par l'indice HHi de contrôle opérationel",
    key="filter_by_hhi_CDV",
)
if filter_by_hhi_CDV:
    hhi_CDV = st.slider(
        "HHI Contrôle opérationel supérieur à :",
        min_value=0.0,
        max_value=1.0,
        value=0.25,
        step=0.01,
        format="%.2f",
    )
    df = df[
        (df["HHI Contrôle opérationel"] >= hhi_CDV)
        | (df["HHI Contrôle opérationel"].isna())
    ]
    l_filtres.append(("HHI Contrôle opérationel supérieur à :", hhi_CDV))

# Filtre hhi_CF
filter_by_hhi_CF = st.checkbox(
    "Filtrer les produits par l'indice HHi de contrôle financier",
    key="filter_by_hhi_CF",
)
if filter_by_hhi_CF:
    hhi_CF = st.slider(
        "HHI Contrôle financier supérieur à :",
        min_value=0.0,
        max_value=1.0,
        value=0.25,
        step=0.01,
        format="%.2f",
    )
    df = df[
        (df["HHI Contrôle financier"] >= hhi_CF)
        | (df["HHI Contrôle financier"].isna())
    ]
    l_filtres.append(("HHI Contrôle financier supérieur à :", hhi_CF))

filter_by_igpc_rank = st.checkbox(
    "Filtrer les produits selon le rang IGPC (score de centralité d'AIPNET)",
    key="filter_by_igpc_rank",
)
if filter_by_igpc_rank:
    igpc_rank = st.slider(
        "Rang IGPC supérieur à :", min_value=0, max_value=100, value=50, step=1
    )
    df = df[(df["rang IGPC (aipnet)"] >= igpc_rank) | (df["rang IGPC (aipnet)"].isna())]
    l_filtres.append(("Rang IGPC supérieur à :", igpc_rank))

# --- Sélection des colonnes à afficher ---
all_columns = df.columns.tolist()
default_columns = df.columns[:15]
selec_columns = st.checkbox("Sélection des colonnes à afficher", key="selec_columns")
if selec_columns:
    selected_columns = st.multiselect(
        "Sélection des colonnes à afficher",
        options=all_columns,
        default=default_columns,
        label_visibility="collapsed",
    )
else:
    selected_columns = default_columns

# --- Affichage du tableau ---
st.write("### Tableau des produits")
if df.empty:
    st.write(
        f"Il n'y a aucun produit (HS6) correspondant aux filtres sélectionnés "
        f"(sur {n_ess_elasticite} produits essentiels selon les critères d'élasticité avant les filtres HHI)."
    )
else:
    st.write(
        f"Il y a {len(df)} produits (HS6) vulnérables, dont "
        f"{df['Essentialité grâce aux élasticités'].sum()} essentiels selon les critères "
        f"d'élasticité (sur {n_ess_elasticite}). De plus, le rang IGPC minimal est "
        f"{int(df['rang IGPC (aipnet)'].min())}, tandis que le rang médian est "
        f"{int(df['rang IGPC (aipnet)'].median())} et qu'il y a "
        f"{(df['rang IGPC (aipnet)'] >= 50).sum()} produits au-dessus de 50."
    )
st.dataframe(df[selected_columns])

# --- Téléchargement des résultats ---
st.write("**Exportations :**")
nom_fichier = st.text_input(
    "Nom du fichier à télécharger (sans extension) :",
    value="resultats_filtres",
    key="nom_fichier",
)
st.download_button(
    label="📥 Télécharger le tableau filtré en csv (sans les métadonnées et informations sur les filtres)",
    data=df.to_csv(index=False, sep=";").encode("utf-8"),
    file_name=f"{nom_fichier}.csv",
    mime="text/csv",
)

if st.button("🔧 Préparer le téléchargement en excel"):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        ws = writer.book.create_sheet("Informations")

        titre_font = Font(size=16, bold=True)
        bold = Font(bold=True)

        ws.merge_cells("A1:B1")
        ws["A1"] = "Productions essentielles"
        ws["A1"].font = titre_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        row = 3

        infos = [
            ("Nombre de produits :", len(df)),
            (
                "Nombres de produits essentiels selon les critères d'élasticité :",
                df["Essentialité grâce aux élasticités"].sum(),
            ),
            (
                "Rang IGPC minimal :",
                int(df["rang IGPC (aipnet)"].min()) if not df.empty else "Non disponible",
            ),
            (
                "Rang IGPC médian :",
                int(df["rang IGPC (aipnet)"].median()) if not df.empty else "Non disponible",
            ),
            (
                "Nombre de produits dont le rang IGPC est supérieur à 50 :",
                (df["rang IGPC (aipnet)"] >= 50).sum(),
            ),
        ]

        for cle, valeur in infos:
            ws[f"A{row}"] = cle
            ws[f"A{row}"].font = bold
            ws[f"B{row}"] = valeur
            row += 1

        row += 1

        if len(l_filtres) > 0:
            ws[f"A{row}"] = "Filtres appliqués"
            ws[f"A{row}"].font = bold
            ws[f"B{row}"] = "Valeur"
            ws[f"B{row}"].font = bold
            row += 1

            for filtre in l_filtres:
                ws[f"A{row}"] = filtre[0]
                ws[f"B{row}"] = filtre[1]
                row += 1
            row += 1

        vert = PatternFill(fill_type="solid", fgColor="C6EFCE")
        orange = PatternFill(fill_type="solid", fgColor="FFF2CC")
        rouge = PatternFill(fill_type="solid", fgColor="F4CCCC")

        ws[f"A{row}"] = "Sections douanières"
        ws[f"A{row}"].font = bold
        ws[f"B{row}"] = "Nombre de catégories HS2 sélectionnées"
        ws[f"B{row}"].font = bold
        ws[f"C{row}"] = "Catégories HS2 sélectionnées"
        ws[f"C{row}"].font = bold
        row += 1
        set_selected = set(selected_categories)
        for section in labels_sections[labels_sections["Niveau"] == "Section"].iterrows():
            ws[f"A{row}"] = "Section " + section[1].Catégorie + " - " + section[1].Label
            l_hs2 = section[1].l_hs2
            ws[f"B{row}"] = f"{len(set_selected & set(l_hs2))} / {len(l_hs2)}"
            ws[f"C{row}"] = ", ".join(sorted(list(set_selected & set(l_hs2))))
            if len(set_selected & set(l_hs2)) == len(l_hs2):
                couleur = vert
            elif len(set_selected & set(l_hs2)) > 0:
                couleur = orange
            else:
                couleur = rouge

            ws[f"A{row}"].fill = couleur
            ws[f"B{row}"].fill = couleur
            ws[f"C{row}"].fill = couleur
            row += 1

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 50
        df.to_excel(writer, sheet_name="Résultats", index=False)
    buffer.seek(0)

    st.download_button(
        label="📥 Télécharger le tableau filtré",
        data=buffer,
        file_name=f"{nom_fichier}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
